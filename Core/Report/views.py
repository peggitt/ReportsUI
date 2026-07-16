from django.shortcuts import render
from django.http import JsonResponse
import json
from django.views.decorators.csrf import csrf_exempt
import datetime
from django.utils import timezone
from django.shortcuts import get_object_or_404
from django.core.exceptions import ObjectDoesNotExist
from django.utils.dateparse import parse_date
from datetime import datetime
from numbers import Number
from Report.models import ReportFile, ReportFileDetail, ReportColumnDefinition
from . import Configs
from openpyxl.utils import get_column_letter



import pymysql
import pandas as pd
from datetime import datetime

# PDF generation
from reportlab.lib.pagesizes import letter, landscape, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

# For visualizations
import matplotlib.pyplot as plt
from io import BytesIO

# For formatting
import re
from django.http import JsonResponse, HttpResponse
import os
from FileProcessing.models import FileProcessSetting
import random
import requests
import zeep

from itertools import groupby
from operator import itemgetter
from operator import attrgetter

SEARCH_POPUP_CONFIG_PATH = os.path.join(os.path.dirname(__file__), "search_popup_config.json")

# Create your views here.
@csrf_exempt
def Home_view(request, *args, **kwargs):
    return render(request,"dashboard.html",{})

@csrf_exempt
def Analytics_view(request, *args, **kwargs):
    return render(request,"analytics.html",{})


@csrf_exempt
def Posted_view(request, *args, **kwargs):
    return render(request,"transactionstatustrue.html",{})

@csrf_exempt
def FailedPosted_view(request, *args, **kwargs):
    return render(request,"transactionstatusfalse.html",{})

@csrf_exempt
def Customers_view(request, *args, **kwargs):

    first_record = FileProcessSetting.objects.first()
    records = []

    if first_record:
        host = first_record.Corehost
        port = first_record.Coreport
        database = first_record.Coredatabase
        username = first_record.Coreusername
        password = first_record.Corepassword
        endpoint_address = first_record.EndpointAddress
        endpoint_address2 = first_record.EndpointAddress2
    

        if Configs.is_ajax(request):
            try:

                hdnReportId = request.POST.get('ActionTag', None)
                IDNo = request.POST.get('IDNo', None)
                print('ActionTag:',hdnReportId,IDNo)
                # Connect to the database
                connection = pymysql.connect(
                    host=host,
                    port=int(port),
                    user=username,
                    password=password,
                    database=database,
                    cursorclass=pymysql.cursors.DictCursor
                )

                if hdnReportId == 'ActivateCustomer':
                    random_code = random.randint(1000, 9999)
                    # Create a cursor and execute the query
                    with connection.cursor() as cursor:
                        sql = """
                        UPDATE tb_mobileaccess 
                        SET isactive = TRUE, password = %s 
                        WHERE IDNo = %s
                        """
                        cursor.execute(sql, (random_code, IDNo))

                        # Commit the changes
                        connection.commit()
                        print("Customer Activated...")

                        
                        
                        #Generate and send the SMS
                        # Select the customer's mobile number
                        sql_select = """
                        SELECT MobileNo,Names,IDNo,DOB FROM tb_mobileaccess 
                        WHERE IDNo = %s
                        """
                        cursor.execute(sql_select, (IDNo,))
                        result = cursor.fetchone()

                        if result:
                            mobile_no = result['MobileNo']

                            names = result['Names']
                            idno = result['IDNo']
                            dob = result['DOB']

                            if mobile_no.startswith('0'):
                                mobile_no = '254' + mobile_no[1:]  # Remove leading '0' and prepend '254'
                            elif not mobile_no.startswith('254'):
                                mobile_no = '254' + mobile_no  # Prepend '254' if it doesn't start with '254'

                            #Call Core and Add the customer
                            wsdl = endpoint_address
                            client = zeep.Client(wsdl=wsdl)

                            clientText = client.service.AddEditLendingCustomer('1003','01',names,'','',idno,mobile_no,dob,'','','CLIENTREGISTRATION')
                            clientTextData = str(clientText)
                            print(clientTextData)

                            
                            urlsend = "https://api.africastalking.com/version1/messaging/bulk"
 
                            payload = json.dumps({
                            "username": "aarcredit",
                            "message": f"Your FinCredit USSD access has been activated. Use the Password : {random_code} to transact.",
                            "senderId": "FINCREDIT",
                            "phoneNumbers": [
                                mobile_no
                            ]
                            })
                            headers = {
                            'Accept': 'application/json',
                            'Content-Type': 'application/json',
                            'apiKey': 'cf59c9adf31d9e3adf674967bd03709292777739981f53dd56da06ee547aab93'
                            }
                            
                            responseSMS = requests.request("POST", urlsend, headers=headers, data=payload)
                            print(responseSMS.text)

                        else:
                            mobile_no = None  # Handle case where customer is not found

                if hdnReportId == 'ResetPassword':
                    random_code = random.randint(1000, 9999)
                    # Create a cursor and execute the query
                    with connection.cursor() as cursor:
                        sql = """
                        UPDATE tb_mobileaccess 
                        SET isactive = TRUE, password = %s 
                        WHERE IDNo = %s
                        """
                        cursor.execute(sql, (random_code, IDNo))

                        # Commit the changes
                        connection.commit()
                        print("Customer Activated...")
                        #Generate and send the SMS
                        # Select the customer's mobile number
                        sql_select = """
                        SELECT MobileNo FROM tb_mobileaccess 
                        WHERE IDNo = %s
                        """
                        cursor.execute(sql_select, (IDNo,))
                        result = cursor.fetchone()

                        if result:
                            mobile_no = result['MobileNo']
                            if mobile_no.startswith('0'):
                                mobile_no = '254' + mobile_no[1:]  # Remove leading '0' and prepend '254'
                            elif not mobile_no.startswith('254'):
                                mobile_no = '254' + mobile_no  # Prepend '254' if it doesn't start with '254'

                            
                            urlsend = "https://api.africastalking.com/version1/messaging/bulk"

                            payload = json.dumps({
                            "username": "aarcredit",
                            "message": f"Your FinCredit USSD access has been activated. Use the Password : {random_code} to transact.",
                            "senderId": "FINCREDIT",
                            "phoneNumbers": [
                                mobile_no
                            ]
                            })
                            headers = {
                            'Accept': 'application/json',
                            'Content-Type': 'application/json',
                            'apiKey': 'cf59c9adf31d9e3adf674967bd03709292777739981f53dd56da06ee547aab93'
                            }
                            
                            responseSMS = requests.request("POST", urlsend, headers=headers, data=payload)
                            print(responseSMS.text)

                        else:
                            mobile_no = None  # Handle case where customer is not found

                response = {'details': 'Action Completed!'}

                return JsonResponse({"Success": response}, status=200)
            
            except  Exception as e:
                eString = str(e)
                print(eString)
                return JsonResponse({"Exception": eString}, status=404)

            finally:
                if 'connection' in locals():
                    connection.close()  # Ensure the connection is closed`
        else:
            try:
                # Connect to the database
                connection = pymysql.connect(
                    host=host,
                    port=int(port),
                    user=username,
                    password=password,
                    database=database,
                    cursorclass=pymysql.cursors.DictCursor
                )

                # Create a cursor and execute the query
                with connection.cursor() as cursor:
                    sql = "SELECT * FROM tb_mobileaccess"
                    cursor.execute(sql)
                    records = cursor.fetchall()  # Fetch all records

            except Exception as e:
                print(f"Error fetching data: {e}")

            finally:
                if 'connection' in locals():
                    connection.close()  # Ensure the connection is closed`

            return render(request,"customers.html",{"Reports": records})

@csrf_exempt
def Reports_view(request, *args, **kwargs):


    try:
        #UserLog = request.session['username']
        #UiD = request.session['Uid']
        #isAdmin = request.session['isAdmin']

        UserLog = "Admin"
        UiD = "Admin"
        isAdmin = True


        if(UserLog=="" or  UiD == ''):
            return render(request,"reports.html",{})
        else:
            view_reports = ReportFile.objects.all()

            # Ensure Reports is sorted by ReportGroup for groupby to work

            Reports = sorted(view_reports, key=attrgetter('ReportGroup'))

           
            # Group the reports by ReportGroup
            grouped_reports = [
                {'grouper': key, 'list': list(group)}
                for key, group in groupby(Reports, key=attrgetter('ReportGroup'))
            ]

            print(grouped_reports)
            return render(request,"reports.html",{'Reports':view_reports,'grouped_reports': grouped_reports})
    except  Exception as e:
            eString = str(e)
            print(eString)
            if eString.find("username") ==1 or eString.find("Uid") ==1:
                return render(request,"reports.html",{})
            else:
                return render(request,"reports.html",{})
            

@csrf_exempt
def ReportDetails_view(request, *args, **kwargs):


    try:
        #UserLog = request.session['username']
        #UiD = request.session['Uid']
        #isAdmin = request.session['isAdmin']

        UserLog = "Admin"
        UiD = "Admin"
        isAdmin = True

        ListItems = list(kwargs.values())
        itemReportId=ListItems[0]


        if(UserLog=="" or  UiD == ''):
            return render(request,"reports.html",{})
        else:
            view_report = ReportFile.objects.filter(ReportId=itemReportId).values().all()
            view_details = ReportFileDetail.objects.filter(ReportId=itemReportId).values().all()
            return render(request,"reportdetails.html",{'Form':view_details,"ReportId":itemReportId,"view_report":view_report})
    except  Exception as e:
            eString = str(e)
            if eString.find("username") ==1 or eString.find("Uid") ==1:
                return render(request,"reports.html",{})
            else:
                return render(request,"reports.html",{})

def construct_parameters(request_data):
    # Ensure request_data is a dictionary
    if isinstance(request_data, str):
        import json
        # Parse the JSON string into a dictionary
        request_data = json.loads(request_data)

    # Define the keys to ignore
    ignore_keys = {"csrfmiddlewaretoken", "hdnReportId", "hdncsrf"}

    # Filter the request_data dictionary to exclude ignored keys
    filtered_params = [
        value for key, value in request_data.items() if key not in ignore_keys
    ]

    return filtered_params

def construct_header(json_data):
    # Initialize the header
    header_parts = []
    # Check if json_data is a string and load it; otherwise, use it directly
    
    if isinstance(json_data, str):
        data = json.loads(json_data)  # Convert JSON string to dictionary
    elif isinstance(json_data, dict):
        data = json_data  # Use the dictionary directly
    else:
        raise ValueError("Input must be a JSON string or a dictionary.")
    # Keys to exclude
    exclude_keys = ["csrfmiddlewaretoken", "hdnReportId", "hdncsrf"]

    # Construct header_parts
    for key, value in data.items():
        if key not in exclude_keys and value:
            header_parts.append(f"{key.replace('ID', '')} {value}")

    # Construct the final report header
    if header_parts:
        Report_Filters = " ".join(header_parts)
        return Report_Filters
    else:
        return ''


def _normalize_search_field_key(value):
    return re.sub(r"[^a-z0-9]", "", str(value or "").lower())


def _load_search_popup_configs():
    with open(SEARCH_POPUP_CONFIG_PATH, "r", encoding="utf-8") as config_file:
        config = json.load(config_file)
    return config.get("fields", {})


@csrf_exempt
def ReportFieldSearchConfig_view(request, *args, **kwargs):
    if request.method != "GET":
        return JsonResponse({"error": "Method not allowed."}, status=405)

    try:
        raw_configs = _load_search_popup_configs()
        response_configs = {}

        for field_name, cfg in raw_configs.items():
            key = _normalize_search_field_key(field_name)
            search_fields = cfg.get("search_fields", [])
            response_configs[key] = {
                "field_name": field_name,
                "title": cfg.get("title", f"Search {field_name}"),
                "search_fields": [
                    {
                        "name": item.get("name"),
                        "label": item.get("label", item.get("name", "")),
                    }
                    for item in search_fields
                    if item.get("name")
                ],
                "display_columns": cfg.get("display_columns", []),
                "value_field": cfg.get("value_field"),
            }

        return JsonResponse({"configs": response_configs}, status=200)
    except Exception as exc:
        return JsonResponse({"error": str(exc)}, status=500)


@csrf_exempt
def ReportFieldSearchData_view(request, *args, **kwargs):
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed."}, status=405)

    connection = None
    try:
        payload = {}
        if request.content_type and "application/json" in request.content_type:
            payload = json.loads(request.body.decode("utf-8") or "{}")
        else:
            payload = request.POST.dict()

        target_field = payload.get("target_field")
        criteria = payload.get("criteria", {})
        if not isinstance(criteria, dict):
            return JsonResponse({"error": "Invalid criteria payload."}, status=400)

        raw_configs = _load_search_popup_configs()
        normalized_target = _normalize_search_field_key(target_field)
        active_config = None
        active_field_name = None

        for field_name, cfg in raw_configs.items():
            if _normalize_search_field_key(field_name) == normalized_target:
                active_config = cfg
                active_field_name = field_name
                break

        if not active_config:
            return JsonResponse({"error": "No search config found for field."}, status=404)

        base_query = str(active_config.get("base_query", "")).strip()
        if not base_query or not base_query.lower().startswith("select"):
            return JsonResponse({"error": "Invalid base query in search config."}, status=500)

        query_parts = [base_query]
        query_params = []

        for search_field in active_config.get("search_fields", []):
            field_name = search_field.get("name")
            condition = search_field.get("condition")
            if not field_name or not condition:
                continue

            value = criteria.get(field_name)
            if value is None:
                value = criteria.get(_normalize_search_field_key(field_name))
            if value is None:
                continue

            value = str(value).strip()
            if not value:
                continue

            match_type = str(search_field.get("match", "contains")).lower()
            if match_type == "contains":
                parameter_value = f"%{value}%"
            elif match_type == "startswith":
                parameter_value = f"{value}%"
            elif match_type == "endswith":
                parameter_value = f"%{value}"
            else:
                parameter_value = value

            query_parts.append(condition)
            query_params.append(parameter_value)

        order_by = str(active_config.get("order_by", "")).strip()
        if order_by:
            query_parts.append(order_by)

        result_limit = active_config.get("result_limit", 100)
        try:
            result_limit = max(1, min(int(result_limit), 500))
        except (TypeError, ValueError):
            result_limit = 100

        final_query = " ".join(query_parts) + " LIMIT %s"
        query_params.append(result_limit)

        first_record = FileProcessSetting.objects.first()
        if not first_record:
            return JsonResponse({"error": "Core database settings not configured."}, status=500)

        connection = connect_to_mysql(
            first_record.Corehost,
            first_record.Coreport,
            first_record.Coredatabase,
            first_record.Coreusername,
            first_record.Corepassword,
        )
        if not connection:
            return JsonResponse({"error": "Unable to connect to core database."}, status=500)

        with connection.cursor() as cursor:
            cursor.execute(final_query, tuple(query_params))
            rows = cursor.fetchall()

        display_columns = active_config.get("display_columns", [])
        if not display_columns and rows:
            display_columns = list(rows[0].keys())

        return JsonResponse(
            {
                "field_name": active_field_name,
                "value_field": active_config.get("value_field"),
                "display_columns": display_columns,
                "rows": rows,
            },
            status=200,
        )
    except Exception as exc:
        return JsonResponse({"error": str(exc)}, status=500)
    finally:
        if connection:
            connection.close()

@csrf_exempt
def RequestReport_view(request, *args, **kwargs):
    if Configs.is_ajax(request):
        try:
            hdnReportId = request.POST.get('hdnReportId', None)
            view_report = ReportFile.objects.filter(ReportId=hdnReportId).values().all()

            for setting in view_report.iterator():
                ReportName = setting['ReportName']         # Accessing using dictionary key
                ReportSDesc = setting['ReportDescription']  # Accessing using dictionary key
                ReportProc = setting['SPObjectName'] 
                

            form_data = {key: request.POST.get(key) for key in request.POST}
            search_json = json.dumps(form_data)
            params = construct_parameters(search_json)

            first_record = FileProcessSetting.objects.first()

            if first_record:
                host = first_record.Corehost
                port = first_record.Coreport
                database = first_record.Coredatabase
                username = first_record.Coreusername
                password = first_record.Corepassword
    


            # MySQL connection details (Edit Accordingly)
            #host = "dev.pathwaystechnologies.com"  # Replace with your MySQL host
            #port = 7878               # Replace with your MySQL port
            #database = "apexbmblive"      # Replace with your database name
            #username = "apexcbs"    # Replace with your username apexcbs/v$Apex@2025
            #password = "v$Apex@2025"    # Replace with your password

            #Connect to database and execute query
            connection = connect_to_mysql(host, port, database, username, password)
            if connection:
                # Execute SQL query
                #df = execute_query(connection, example_query)
                
                # Alternative: Execute a stored procedure
                df,dfHeaders = execute_stored_procedure(connection, ReportProc,params)
                print('Executing Report')
                now = datetime.now()

                # Format the datetime string
                formatted_datetime = now.strftime("%Y%m%d_%H%M%S")

                # Generate PDF report
                if df.empty:
                    raise ValueError("No data returned for the selected filters.")

                create_pdf(
                    df, 
                    dfHeaders,
                    f"report_{formatted_datetime}.pdf", 
                    title=ReportName,
                    subtitle=ReportSDesc,
                    include_chart=False,
                    report_id=hdnReportId
                )
                print('Executing Report Completed')
                # Close connection
            connection.close()
            fileName = f"report_{formatted_datetime}.pdf"

            response = {
                'details': 'Report Generated!',
                'file_url': request.build_absolute_uri(f'/media/ReportFiles/{fileName}')  # URL to download the file
            }
            
            return JsonResponse({"Success": response}, status=200)
        except  Exception as e:
            eString = str(e)
            print(eString)
            return JsonResponse({"Exception": eString}, status=404)


def download_report(request, file_name):
    file_path = os.path.join('media/ReportFiles/', file_name)
    
    if os.path.exists(file_path):
        with open(file_path, 'rb') as f:
            if file_name.endswith('.pdf'):
                content_type = 'application/pdf'
            elif file_name.endswith('.xlsx'):
                content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            else:
                return JsonResponse({"error": "Unsupported file type"}, status=400)

            response = HttpResponse(f.read(), content_type=content_type)
            response['Content-Disposition'] = f'attachment; filename="{file_name}"'
            return response
    else:
        return JsonResponse({"error": "File not found"}, status=404)
    
def connect_to_mysql(host, port, database, username, password):
    """
    Establish a connection to MySQL database
    
    Returns:
        connection: MySQL connection object
    """
    try:
        print('Lets Connect:'+host,port,database,username,password)
        connection = pymysql.connect(
            host=host,
            port=int(port),
            user=username,
            password=password,
            database=database,
            cursorclass=pymysql.cursors.DictCursor,
            ssl=False  # Disable SSL
        )
        print(f"Successfully connected to MySQL database: {database} at {host}:{port}")
        return connection
    except Exception as e:
        print(f"Error connecting to database: {e} : {database} at {host}:{port}")
        return None


def execute_query(connection, query, params=None):
    """
    Execute a SQL query and return the results
    
    Args:
        connection: MySQL connection object
        query: SQL query string
        params: Parameters for the query (optional)
        
    Returns:
        pandas.DataFrame: Results as a DataFrame
    """
    try:
        with connection.cursor() as cursor:
            cursor.execute(query, params or ())
            results = cursor.fetchall()
            return pd.DataFrame(results)
    except Exception as e:
        print(f"Error executing query: {e}")
        return pd.DataFrame()

    
def execute_stored_procedure(connection,procedure_name, params=None):
    """
    Execute a stored procedure that may return multiple result sets
    Returns two separate DataFrames (df1, df2)
    If second result set doesn't exist, returns empty DataFrame for df2
    """
    param_placeholders = ', '.join(['%s'] * len(params)) if params else ''
    query = f"CALL {procedure_name}({param_placeholders})"
    
    try:
        with connection.cursor() as cursor:
            cursor.execute(query, params or ())
            
            # Get first result set
            df1 = pd.DataFrame()
            df2 = pd.DataFrame()
            
            try:
                rows = cursor.fetchall()
                if rows:
                    columns = [col[0] for col in cursor.description]
                    df1 = pd.DataFrame(rows, columns=columns)
                    print(f"First result set shape: {df1.shape}")
                else:
                    print("First result set is empty")
            except Exception as e:
                print(f"Error fetching first result set: {e}")
                df1 = pd.DataFrame()
            
            # Try to get second result set
            try:
                if cursor.nextset():
                    rows = cursor.fetchall()
                    if rows:
                        columns = [col[0] for col in cursor.description]
                        df2 = pd.DataFrame(rows, columns=columns)
                        print(f"Second result set shape: {df2.shape}")
                    else:
                        print("Second result set is empty")
                else:
                    print("No second result set available")
            except Exception as e:
                # Handle case where no more result sets exist
                if "no result set" in str(e).lower() or "no data" in str(e).lower():
                    print("No second result set available")
                else:
                    print(f"Error fetching second result set: {e}")
                df2 = pd.DataFrame()
            
            return df1, df2
            
    except Exception as e:
        error_message = f"Error executing stored procedure {procedure_name}: {e}"
        print(error_message)
        raise ValueError(error_message)
    
    
PDF_PRIMARY = colors.HexColor("#102C57")
PDF_PRIMARY_SOFT = colors.HexColor("#173C73")
PDF_ACCENT = colors.HexColor("#1EA7B8")
PDF_BORDER = colors.HexColor("#C7D5EB")
PDF_ROW_ALT = colors.HexColor("#F4F8FE")
PDF_TEXT = colors.HexColor("#1A2E4E")
PDF_TEXT_MUTED = colors.HexColor("#5A6D8A")


def _select_page_size(page_size, orientation):
    page_size_tuple = A4 if page_size == 'A4' else letter
    if orientation == 'landscape':
        page_size_tuple = landscape(page_size_tuple)
    return page_size_tuple


def _safe_logo_path(logo_path=None):
    candidates = []
    if logo_path:
        candidates.append(logo_path)
    candidates.extend(["Fincredit.png", "LOGO-BONAFIDE.png"])

    for candidate in candidates:
        if candidate and os.path.exists(candidate):
            return candidate
    return None


def _build_pdf_styles():
    base_styles = getSampleStyleSheet()
    return {
        'title': ParagraphStyle(
            'CorporateTitle',
            parent=base_styles['Heading1'],
            fontName='Helvetica-Bold',
            fontSize=18,
            leading=22,
            alignment=1,
            textColor=PDF_PRIMARY,
            spaceAfter=6
        ),
        'subtitle': ParagraphStyle(
            'CorporateSubtitle',
            parent=base_styles['Heading3'],
            fontName='Helvetica-Bold',
            fontSize=10,
            leading=13,
            alignment=1,
            textColor=PDF_PRIMARY_SOFT,
            spaceAfter=4
        ),
        'meta': ParagraphStyle(
            'CorporateMeta',
            parent=base_styles['Normal'],
            fontName='Helvetica',
            fontSize=8,
            leading=10,
            alignment=1,
            textColor=PDF_TEXT_MUTED
        ),
        'normal': ParagraphStyle(
            'CorporateNormal',
            parent=base_styles['Normal'],
            fontName='Helvetica',
            fontSize=8,
            leading=10,
            textColor=PDF_TEXT
        ),
        'bold': ParagraphStyle(
            'CorporateBold',
            parent=base_styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=8,
            leading=10,
            textColor=PDF_TEXT
        ),
        'table_header': ParagraphStyle(
            'CorporateTableHeader',
            parent=base_styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=8,
            leading=10,
            alignment=1,
            textColor=colors.white
        )
    }


def _append_report_header(elements, doc, styles, title, subtitle=None, logo_path=None):
    resolved_logo = _safe_logo_path(logo_path)
    if resolved_logo:
        logo = Image(resolved_logo, width=3.8 * inch, height=1.0 * inch)
        logo.hAlign = 'CENTER'
        elements.append(logo)
        elements.append(Spacer(1, 0.1 * inch))

    elements.append(Paragraph(title, styles['title']))
    if subtitle:
        elements.append(Paragraph(subtitle, styles['subtitle']))

    generated_label = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    elements.append(Paragraph(generated_label, styles['meta']))
    elements.append(Spacer(1, 0.1 * inch))

    title_bar = Table([[""]], colWidths=[doc.width], rowHeights=[0.08 * inch])
    title_bar.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), PDF_ACCENT),
        ('LINEBELOW', (0, 0), (-1, -1), 0.5, PDF_ACCENT),
    ]))
    elements.append(title_bar)
    elements.append(Spacer(1, 0.16 * inch))


def _coerce_value(value):
    if pd.isna(value):
        return ""
    return str(value)


def _normalize_col(name):
    return str(name).strip().lower()


def _is_numeric_type_name(data_type):
    if not data_type:
        return False
    numeric_tokens = {"number", "numeric", "int", "integer", "decimal", "float", "double", "money", "amount"}
    return any(token in str(data_type).strip().lower() for token in numeric_tokens)


def _is_numeric_column(series, data_type=None):
    if _is_numeric_type_name(data_type):
        return True
    return pd.api.types.is_numeric_dtype(series)


def _format_report_value(value, is_numeric=False):
    if value is None or pd.isna(value):
        return ""
    if is_numeric and isinstance(value, Number) and not isinstance(value, bool):
        if isinstance(value, float) and not float(value).is_integer():
            return f"{value:,.2f}"
        return f"{int(value):,}"
    return str(value)


def _to_pdf_col_width(raw_width):
    if raw_width is None:
        return None
    try:
        width = float(raw_width)
        if width <= 0:
            return None
        return max(0.8, min(6.0, width * 0.12)) * inch
    except (TypeError, ValueError):
        return None


def _to_excel_col_width(raw_width):
    if raw_width is None:
        return None
    try:
        width = float(raw_width)
        if width <= 0:
            return None
        return width
    except (TypeError, ValueError):
        return None


def get_report_column_config(report_id, df):
    if df.empty:
        return [], {}, {}, {}, {}

    data_columns = [col for col in df.columns if _normalize_col(col) != "descr"]
    defs = ReportColumnDefinition.objects.filter(ReportId=report_id).order_by("SortOrder", "ReportColumnId")

    if not defs.exists():
        default_numeric = {
            col: _is_numeric_column(df[col])
            for col in data_columns
        }
        default_align = {col: "LEFT" for col in data_columns}
        return data_columns, {col: col for col in data_columns}, {}, default_numeric, default_align

    df_cols_by_norm = {_normalize_col(col): col for col in data_columns}
    selected_columns = []
    display_names = {}
    column_widths = {}
    numeric_columns = {}
    column_alignments = {}

    for col_def in defs:
        normalized_name = _normalize_col(col_def.ColumnName)
        source_col = df_cols_by_norm.get(normalized_name)
        if not source_col:
            continue
        if not col_def.isVisible:
            continue
        if source_col not in selected_columns:
            selected_columns.append(source_col)
        display_names[source_col] = (col_def.DisplayName or source_col).strip()
        column_widths[source_col] = col_def.ColumnWidth
        numeric_columns[source_col] = _is_numeric_column(df[source_col], col_def.DataType)
        column_alignments[source_col] = (col_def.LoadSide or "LEFT").upper()

    if not selected_columns:
        selected_columns = data_columns
        for col in selected_columns:
            display_names[col] = col
            numeric_columns[col] = _is_numeric_column(df[col])
            column_alignments[col] = "LEFT"

    return selected_columns, display_names, column_widths, numeric_columns, column_alignments


def _build_filter_table(df_headers, styles, available_width):
    if df_headers.empty:
        return None

    rows = []
    for _, row in df_headers.iterrows():
        for key, value in row.items():
            key_text = _coerce_value(key).strip()
            value_text = _coerce_value(value).strip()
            if key_text and value_text:
                rows.append([
                    Paragraph(f"<b>{key_text}</b>", styles['bold']),
                    Paragraph(value_text, styles['normal'])
                ])

    if not rows:
        return None

    label_width = min(2.2 * inch, available_width * 0.32)
    value_width = max(available_width - label_width, 2.2 * inch)
    table = Table(rows, colWidths=[label_width, value_width])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor("#EEF4FF")),
        ('TEXTCOLOR', (0, 0), (-1, -1), PDF_TEXT),
        ('GRID', (0, 0), (-1, -1), 0.35, PDF_BORDER),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    return table


def _default_col_widths(doc_width, column_count):
    if column_count <= 0:
        return []
    return [doc_width / column_count] * column_count


def create_pdf_Logo(df, output_path, title="Database Report", subtitle=None,
               orientation='portrait', page_size='A4', include_chart=False, col_widths=None, logo_path=None):
    """
    Generate a branded PDF report from a pandas DataFrame.
    """
    doc = SimpleDocTemplate(
        os.path.join("media", "ReportFiles", output_path),
        pagesize=_select_page_size(page_size, orientation),
        rightMargin=0.5 * inch,
        leftMargin=0.5 * inch,
        topMargin=0.45 * inch,
        bottomMargin=0.45 * inch
    )

    styles = _build_pdf_styles()
    elements = []
    _append_report_header(elements, doc, styles, title=title, subtitle=subtitle, logo_path=logo_path)

    if include_chart and not df.empty and df.shape[0] > 0 and df.shape[1] > 1:
        try:
            numeric_cols = df.select_dtypes(include=['number']).columns
            if len(numeric_cols) > 0:
                plt.figure(figsize=(7, 3.6))
                if df.shape[0] <= 10:
                    df.plot(x=df.columns[0], y=numeric_cols[0], kind='bar', legend=False, color='#1E5A96')
                else:
                    df.plot(x=df.columns[0], y=numeric_cols[0], kind='line', legend=False, color='#1E5A96')
                plt.title(f"{numeric_cols[0]} Trend", fontsize=11, color='#102C57')
                plt.tight_layout()
                img_data = BytesIO()
                plt.savefig(img_data, format='png', dpi=150)
                plt.close()
                img_data.seek(0)
                chart_img = Image(img_data, width=5.9 * inch, height=2.4 * inch)
                chart_img.hAlign = 'CENTER'
                elements.append(chart_img)
                elements.append(Spacer(1, 0.12 * inch))
        except Exception as e:
            print(f"Error creating chart: {e}")

    if not df.empty:
        header_row = [Paragraph(_coerce_value(col), styles['table_header']) for col in df.columns]
        body_rows = [
            [Paragraph(_coerce_value(item), styles['normal']) for item in row]
            for row in df.values.tolist()
        ]
        data = [header_row] + body_rows
        resolved_col_widths = col_widths if col_widths else _default_col_widths(doc.width, len(df.columns))
        table = Table(data, colWidths=resolved_col_widths, repeatRows=1)
        apply_table_style(table)
        elements.append(table)
    else:
        elements.append(Paragraph("No data available.", styles['normal']))

    doc.build(elements)
    print(f"PDF report saved to: {output_path}")
    return output_path


def create_pdf(df, dfHeaders, output_path, title="Database Report", subtitle=None,
               orientation='landscape', page_size='A4', include_chart=False, report_id=None):
    """
    Generate a branded PDF report from a pandas DataFrame.
    """
    doc = SimpleDocTemplate(
        os.path.join("media", "ReportFiles", output_path),
        pagesize=_select_page_size(page_size, orientation),
        rightMargin=0.5 * inch,
        leftMargin=0.5 * inch,
        topMargin=0.45 * inch,
        bottomMargin=0.45 * inch
    )

    styles = _build_pdf_styles()
    elements = []
    _append_report_header(elements, doc, styles, title=title, subtitle=subtitle, logo_path="Fincredit.png")

    header_table = _build_filter_table(dfHeaders, styles, doc.width)
    if header_table:
        elements.append(header_table)
        elements.append(Spacer(1, 0.16 * inch))

    if include_chart and not df.empty and df.shape[0] > 0 and df.shape[1] > 1:
        try:
            numeric_cols = df.select_dtypes(include=['number']).columns
            if len(numeric_cols) > 0:
                plt.figure(figsize=(7, 3.4))
                if df.shape[0] <= 10:
                    df.plot(x=df.columns[0], y=numeric_cols[0], kind='bar', legend=False, color='#1E5A96')
                else:
                    df.plot(x=df.columns[0], y=numeric_cols[0], kind='line', legend=False, color='#1E5A96')
                plt.title(f"{numeric_cols[0]} Trend", fontsize=11, color='#102C57')
                plt.tight_layout()
                img_data = BytesIO()
                plt.savefig(img_data, format='png', dpi=150)
                plt.close()
                img_data.seek(0)
                chart_img = Image(img_data, width=6.7 * inch, height=2.3 * inch)
                chart_img.hAlign = 'CENTER'
                elements.append(chart_img)
                elements.append(Spacer(1, 0.14 * inch))
        except Exception as e:
            print(f"Error creating chart: {e}")

    if df.empty:
        elements.append(Paragraph("No data available.", styles['normal']))
        doc.build(elements)
        print(f"PDF report saved to: {output_path}")
        return output_path

    selected_columns, display_names, column_widths, numeric_columns, column_alignments = get_report_column_config(report_id, df)
    if not selected_columns:
        elements.append(Paragraph("No visible report columns to display.", styles['normal']))
        doc.build(elements)
        print(f"PDF report saved to: {output_path}")
        return output_path

    table_header = [Paragraph(_coerce_value(display_names.get(col, col)), styles['table_header']) for col in selected_columns]
    table_rows = []
    default_col_width = doc.width / len(selected_columns)
    col_widths = []
    for col in selected_columns:
        configured_width = _to_pdf_col_width(column_widths.get(col))
        col_widths.append(configured_width or default_col_width)
    right_aligned_col_indices = []
    for idx, col in enumerate(selected_columns):
        configured_align = column_alignments.get(col, "LEFT")
        if configured_align == "RIGHT" or numeric_columns.get(col, False):
            right_aligned_col_indices.append(idx)
    has_data_table = False

    def flush_rows(add_page_break=False):
        nonlocal has_data_table
        if table_rows:
            table = Table([table_header] + table_rows, colWidths=col_widths, repeatRows=1)
            apply_table_style(table, right_aligned_col_indices=right_aligned_col_indices)
            elements.append(table)
            table_rows.clear()
            has_data_table = True
        if add_page_break:
            elements.append(PageBreak())

    for _, row in df.iterrows():
        descriptor = _coerce_value(row.get('descr', '')).strip().lower() if 'descr' in row else ''

        if descriptor == 'pagebreak':
            flush_rows(add_page_break=True)
            continue

        row_style = styles['bold'] if descriptor == 'bold' else styles['normal']
        formatted_row = [
            Paragraph(_format_report_value(row.get(col, ''), numeric_columns.get(col, False)), row_style)
            for col in selected_columns
        ]
        table_rows.append(formatted_row)

    flush_rows(add_page_break=False)
    if not has_data_table:
        elements.append(Paragraph("No data to display after processing.", styles['normal']))

    doc.build(elements)
    print(f"PDF report saved to: {output_path}")
    return output_path
def format_output(value):
    if isinstance(value, str):
        return f"String: {value}"
    elif isinstance(value, int):
        return f"Integer: {value:,}"  # Comma as thousands separator
    elif isinstance(value, float):
        return f"Float: {value:.2f}"  # 2 decimal places
    elif isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    else:
        return "Unsupported type"

def apply_table_style(table, right_aligned_col_indices=None):
    """Apply corporate banking styling to report tables."""
    table_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), PDF_PRIMARY),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 6),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 7),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, PDF_ROW_ALT]),
        ('TEXTCOLOR', (0, 1), (-1, -1), PDF_TEXT),
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 7.5),
        ('GRID', (0, 0), (-1, -1), 0.35, PDF_BORDER),
        ('LINEABOVE', (0, 0), (-1, 0), 0.4, PDF_PRIMARY_SOFT),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
    ])
    if right_aligned_col_indices:
        for col_idx in right_aligned_col_indices:
            table_style.add('ALIGN', (col_idx, 1), (col_idx, -1), 'RIGHT')
    table.setStyle(table_style)


def parse_formatting_tags(text):
    """
    Parse formatting tags in text and return formatted result
    
    Supported tags:
    - [bold]text[/bold] - Makes text bold
    - [header]text[/header] - Makes text a header
    - [pagebreak] - Inserts a page break
    
    Args:
        text: Text with formatting tags
        
    Returns:
        List of elements with appropriate formatting
    """
    corporate_styles = _build_pdf_styles()
    normal_style = corporate_styles['normal']
    heading_style = ParagraphStyle(
        'CorporateHeading',
        parent=corporate_styles['bold'],
        fontSize=11,
        leading=14,
        alignment=1,
        textColor=PDF_PRIMARY
    )
    bold_style = corporate_styles['bold']
    
    elements = []
    
    # Split by pagebreak tag
    parts = text.split('[pagebreak]')
    
    for i, part in enumerate(parts):
        # Process bold tags
        bold_parts = re.split(r'\[bold\](.*?)\[\/bold\]', part)
        
        for j, bold_part in enumerate(bold_parts):
            if j % 2 == 0:  # Regular text
                # Process header tags
                header_parts = re.split(r'\[header\](.*?)\[\/header\]', bold_part)
                
                for k, header_part in enumerate(header_parts):
                    if k % 2 == 0:  # Regular text
                        if header_part.strip():
                            elements.append(Paragraph(header_part, normal_style))
                    else:  # Header text
                        elements.append(Paragraph(header_part, heading_style))
            else:  # Bold text
                elements.append(Paragraph(bold_part, bold_style))
        
        # Add page break if not the last part
        if i < len(parts) - 1:
            elements.append(PageBreak())
    
    return elements

def create_formatted_pdf(formatted_text, output_path, title="Formatted Report", 
                        orientation='portrait', page_size='A4'):
    """
    Generate a PDF with formatted text
    
    Args:
        formatted_text: Text with formatting tags
        output_path: Path where PDF will be saved
        title: Report title
        orientation: 'portrait' or 'landscape'
        page_size: 'A4' or 'letter'
    """
    # Set up PDF document properties
    if page_size == 'A4':
        page_size_tuple = A4
    else:  # letter
        page_size_tuple = letter
        
    if orientation == 'landscape':
        page_size_tuple = landscape(page_size_tuple)
    
    doc = SimpleDocTemplate(
        output_path, 
        pagesize=page_size_tuple,
        rightMargin=0.5*inch,
        leftMargin=0.5*inch,
        topMargin=0.5*inch,
        bottomMargin=0.5*inch
    )
    
    styles = _build_pdf_styles()
    
    # Create document elements
    elements = []
    
    # Add title
    elements.append(Paragraph(title, styles['title']))
    elements.append(Spacer(1, 0.18 * inch))
    
    # Add formatted content
    elements.extend(parse_formatting_tags(formatted_text))
    
    # Build PDF
    doc.build(elements)
    print(f"Formatted PDF report saved to: {output_path}")
    return output_path

@csrf_exempt
def RequestReportxlsx_view(request, *args, **kwargs):

    from openpyxl import Workbook
    from openpyxl.drawing.image import Image
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import Alignment, Font

    if Configs.is_ajax(request):
        try:
            hdnReportId = request.POST.get('hdnReportId', None)
            view_report = ReportFile.objects.filter(ReportId=hdnReportId).values().all()

            for setting in view_report.iterator():
                ReportName = setting['ReportName']         # Accessing using dictionary key
                ReportSDesc = setting['ReportDescription']  # Accessing using dictionary key
                ReportProc = setting['SPObjectName'] 
                

            form_data = {key: request.POST.get(key) for key in request.POST}
            search_json = json.dumps(form_data)
            params = construct_parameters(search_json)
            print(params)
            
            first_record = FileProcessSetting.objects.first()

            if first_record:
                host = first_record.Corehost
                port = first_record.Coreport
                database = first_record.Coredatabase
                username = first_record.Coreusername
                password = first_record.Corepassword
                
            # MySQL connection details (Edit Accordingly)
            #host = "dev.pathwaystechnologies.com"  # Replace with your MySQL host
            #port = 7878               # Replace with your MySQL port
            #database = "apexbmblive"      # Replace with your database name
            #username = "apexcbs"    # Replace with your username apexcbs/v$Apex@2025
            #password = "v$Apex@2025"    # Replace with your password

            #Connect to database and execute query
            connection = connect_to_mysql(host, port, database, username, password)
            if connection:
                # Execute SQL query
                #df = execute_query(connection, example_query)
                
                # Alternative: Execute a stored procedure
                df,dfheaders = execute_stored_procedure(connection, ReportProc,params)
                #print(df)
                #print(dfheaders)
               # if 'descr' in df.columns:
               #     df = df.drop(columns=['descr'])
                    
                print('Executing Report')
                now = datetime.now()

                # Format the datetime string
                formatted_datetime = now.strftime("%Y%m%d_%H%M%S")
                

                # Generate EXCEL report
                if df.empty:
                    raise ValueError("No data returned for the selected filters.")

                selected_columns, display_names, column_widths, numeric_columns, column_alignments = get_report_column_config(hdnReportId, df)
                excel_df = df[selected_columns].copy() if selected_columns else df.copy()
                excel_df.rename(columns=display_names, inplace=True)

                # Create a new workbook
                wb = Workbook()
                ws = wb.active
                    
                # Add a logo (replace 'logo.png' with your logo path)
                logo = Image("Fincredit.png")  # Ensure the logo file exists
                logo.height = 50  # Adjust height as needed
                logo.width = 350  # Adjust width as needed
                ws.add_image(logo, 'D1')  # Place logo in cell A1

                # Add report title below the logo (row 2)
                report_title = ReportName  # Replace with your title
                ws['A5'] = report_title
                ws['A5'].font = Font(size=16, bold=True)
                ws.merge_cells('A5:H5')  # Merge cells for the title (adjust range as needed)
                ws['A5'].alignment = Alignment(horizontal='center')
               # for cell in ws[8]:
               #     cell.font = Font(size=14, bold=True)
                print(' headers')
                Report_Filters = construct_header(search_json)

                print('end headers')
                #Report_Filters ="Report For Branch " + params[0] + " To Branch " + params[1] + " From " + params[2] + " To " + params[3]
                #ws['A9'].font = Font(size=14, bold=True) 
                #ws.merge_cells('A9:C9')
                ws['B6'].alignment = Alignment(horizontal='center') 
                ws['B6'] = Report_Filters
                ws.merge_cells('A6:D6')
                print(Report_Filters)
                for row in dataframe_to_rows(excel_df, index=False, header=True):
                    ws.append(row)

                for cell in ws[7]:
                    cell.font = Font(size=14, bold=True)

                for row_idx in range(8, ws.max_row + 1):
                    for col_idx, source_col in enumerate(selected_columns, 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        value = cell.value
                        if value is None:
                            continue

                        force_right = column_alignments.get(source_col, "LEFT") == "RIGHT"
                        is_numeric = numeric_columns.get(source_col, False)
                        if is_numeric and isinstance(value, Number) and not isinstance(value, bool):
                            if isinstance(value, float) and not float(value).is_integer():
                                cell.number_format = '#,##0.00'
                            else:
                                cell.number_format = '#,##0'
                            cell.alignment = Alignment(horizontal='right')
                        else:
                            cell.alignment = Alignment(horizontal='right' if force_right else 'left')

                if ws.max_row >= 8:
                    for cell in ws[ws.max_row]:
                        cell.font = Font(size=11, bold=False)

                for col_idx, source_col in enumerate(selected_columns, 1):
                    column_letter = get_column_letter(col_idx)
                    configured_width = _to_excel_col_width(column_widths.get(source_col))
                    if configured_width:
                        ws.column_dimensions[column_letter].width = configured_width
                        continue

                    max_length = len(str(display_names.get(source_col, source_col)))
                    for row_idx in range(8, ws.max_row + 1):
                        value = ws.cell(row=row_idx, column=col_idx).value
                        if value is None:
                            continue
                        max_length = max(max_length, len(str(value)))
                    ws.column_dimensions[column_letter].width = max_length + 2
                    
                print("df here")    
                # Save the file
                output_path = os.path.join("media/ReportFiles", f"report_{formatted_datetime}.xlsx")  # Saves as 'media/report.xlsx'
                wb.save(output_path)
                print("done")
                print('Executing Report Completed')
                # Close connection
            connection.close()
            fileName = f"report_{formatted_datetime}.xlsx"

            response = {
                'details': 'Report Generated!',
                'file_url': request.build_absolute_uri(f'/media/ReportFiles/{fileName}')  # URL to download the file
            }
            
            return JsonResponse({"Success": response}, status=200)
        except  Exception as e:
            eString = str(e)
            print(eString)
            return JsonResponse({"Exception": eString}, status=404)
