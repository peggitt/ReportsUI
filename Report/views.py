from django.shortcuts import render
from django.http import JsonResponse
import json
from django.views.decorators.csrf import csrf_exempt
import datetime
from django.utils import timezone
from django.shortcuts import get_object_or_404
from django.core.exceptions import ObjectDoesNotExist
from django.utils.dateparse import parse_date
from datetime import datetime
from Report.models import ReportFile,ReportFileDetail
from . import Configs
from openpyxl.utils import get_column_letter



import pymysql
import pandas as pd
from datetime import datetime

# PDF generation
from reportlab.lib.pagesizes import letter, landscape, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

# For visualizations
import matplotlib.pyplot as plt
from io import BytesIO

# For formatting
import re
from django.http import JsonResponse, HttpResponse
import os

# Create your views here.
@csrf_exempt
def Home_view(request, *args, **kwargs):
    return render(request,"dashboard.html",{})

@csrf_exempt
def Analytics_view(request, *args, **kwargs):
    return render(request,"analytics.html",{})

@csrf_exempt
def Reports_view(request, *args, **kwargs):


    try:
        #UserLog = request.session['username']
        #UiD = request.session['Uid']
        #isAdmin = request.session['isAdmin']

        UserLog = "Admin"
        UiD = "Admin"
        isAdmin = True


        if(UserLog=="" or  UiD == ''):
            return render(request,"reports.html",{})
        else:
            view_reports = ReportFile.objects.all()
            
            return render(request,"reports.html",{'Reports':view_reports})
    except  Exception as e:
            eString = str(e)
            print(eString)
            if eString.find("username") ==1 or eString.find("Uid") ==1:
                return render(request,"reports.html",{})
            else:
                return render(request,"reports.html",{})
            

@csrf_exempt
def ReportDetails_view(request, *args, **kwargs):


    try:
        #UserLog = request.session['username']
        #UiD = request.session['Uid']
        #isAdmin = request.session['isAdmin']

        UserLog = "Admin"
        UiD = "Admin"
        isAdmin = True

        ListItems = list(kwargs.values())
        itemReportId=ListItems[0]


        if(UserLog=="" or  UiD == ''):
            return render(request,"reports.html",{})
        else:
            view_report = ReportFile.objects.filter(ReportId=itemReportId).values().all()
            view_details = ReportFileDetail.objects.filter(ReportId=itemReportId).values().all()
            return render(request,"reportdetails.html",{'Form':view_details,"ReportId":itemReportId,"view_report":view_report})
    except  Exception as e:
            eString = str(e)
            if eString.find("username") ==1 or eString.find("Uid") ==1:
                return render(request,"reports.html",{})
            else:
                return render(request,"reports.html",{})

def construct_parameters(request_data):
    # Ensure request_data is a dictionary
    if isinstance(request_data, str):
        import json
        # Parse the JSON string into a dictionary
        request_data = json.loads(request_data)

    # Define the keys to ignore
    ignore_keys = {"csrfmiddlewaretoken", "hdnReportId", "hdncsrf"}

    # Filter the request_data dictionary to exclude ignored keys
    filtered_params = [
        value for key, value in request_data.items() if key not in ignore_keys
    ]

    return filtered_params

def construct_header(json_data):
    # Initialize the header
    header_parts = []
    # Check if json_data is a string and load it; otherwise, use it directly
    
    if isinstance(json_data, str):
        data = json.loads(json_data)  # Convert JSON string to dictionary
    elif isinstance(json_data, dict):
        data = json_data  # Use the dictionary directly
    else:
        raise ValueError("Input must be a JSON string or a dictionary.")
    # Keys to exclude
    exclude_keys = ["csrfmiddlewaretoken", "hdnReportId", "hdncsrf"]

    # Construct header_parts
    for key, value in data.items():
        if key not in exclude_keys and value:
            header_parts.append(f"{key.replace('ID', '')} {value}")

    # Construct the final report header
    if header_parts:
        Report_Filters = " ".join(header_parts)
        return Report_Filters
    else:
        return ''

@csrf_exempt
def RequestReport_view(request, *args, **kwargs):
    if Configs.is_ajax(request):
        try:
            hdnReportId = request.POST.get('hdnReportId', None)
            view_report = ReportFile.objects.filter(ReportId=hdnReportId).values().all()

            for setting in view_report.iterator():
                ReportName = setting['ReportName']         # Accessing using dictionary key
                ReportSDesc = setting['ReportDescription']  # Accessing using dictionary key
                ReportProc = setting['SPObjectName'] 
                

            form_data = {key: request.POST.get(key) for key in request.POST}
            search_json = json.dumps(form_data)
            params = construct_parameters(search_json)

            # MySQL connection details (Edit Accordingly)
            host = "localhost"  # Replace with your MySQL host
            port = 3306               # Replace with your MySQL port
            database = "fincredit"      # Replace with your database name
            username = "root"    # Replace with your username
            password = "Pass1"    # Replace with your password


            # MySQL connection details (Edit Accordingly)
            #host = "dev.pathwaystechnologies.com"  # Replace with your MySQL host
            #port = 7878               # Replace with your MySQL port
            #database = "apexbmblive"      # Replace with your database name
            #username = "apexcbs"    # Replace with your username apexcbs/v$Apex@2025
            #password = "v$Apex@2025"    # Replace with your password

            #Connect to database and execute query
            connection = connect_to_mysql(host, port, database, username, password)
            if connection:
                # Execute SQL query
                #df = execute_query(connection, example_query)
                
                # Alternative: Execute a stored procedure
                df,dfHeaders = execute_stored_procedure(connection, ReportProc,params)
                print('Executing Report')
                now = datetime.now()

                # Format the datetime string
                formatted_datetime = now.strftime("%Y%m%d_%H%M%S")

                # Generate PDF report
                if not df.empty:
                    create_pdf(
                        df, 
                        dfHeaders,
                        f"report_{formatted_datetime}.pdf", 
                        title=ReportName,
                        subtitle=ReportSDesc,
                        include_chart=False
                    )
                print('Executing Report Completed')
                # Close connection
            connection.close()
            fileName = f"report_{formatted_datetime}.pdf"

            response = {
                'details': 'Report Generated!',
                'file_url': request.build_absolute_uri(f'/media/ReportFiles/{fileName}')  # URL to download the file
            }
            
            return JsonResponse({"Success": response}, status=200)
        except  Exception as e:
            eString = str(e)
            print(eString)
            return JsonResponse({"Exception": eString}, status=404)


def download_report(request, file_name):
    file_path = os.path.join('media/ReportFiles/', file_name)
    
    if os.path.exists(file_path):
        with open(file_path, 'rb') as f:
            if file_name.endswith('.pdf'):
                content_type = 'application/pdf'
            elif file_name.endswith('.xlsx'):
                content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            else:
                return JsonResponse({"error": "Unsupported file type"}, status=400)

            response = HttpResponse(f.read(), content_type=content_type)
            response['Content-Disposition'] = f'attachment; filename="{file_name}"'
            return response
    else:
        return JsonResponse({"error": "File not found"}, status=404)
    
def connect_to_mysql(host, port, database, username, password):
    """
    Establish a connection to MySQL database
    
    Returns:
        connection: MySQL connection object
    """
    try:
        connection = pymysql.connect(
            host=host,
            port=int(port),
            user=username,
            password=password,
            database=database,
            cursorclass=pymysql.cursors.DictCursor
        )
        print(f"Successfully connected to MySQL database: {database} at {host}:{port}")
        return connection
    except Exception as e:
        print(f"Error connecting to database: {e}")
        return None


def execute_query(connection, query, params=None):
    """
    Execute a SQL query and return the results
    
    Args:
        connection: MySQL connection object
        query: SQL query string
        params: Parameters for the query (optional)
        
    Returns:
        pandas.DataFrame: Results as a DataFrame
    """
    try:
        with connection.cursor() as cursor:
            cursor.execute(query, params or ())
            results = cursor.fetchall()
            return pd.DataFrame(results)
    except Exception as e:
        print(f"Error executing query: {e}")
        return pd.DataFrame()

    
def execute_stored_procedure(connection,procedure_name, params=None):
    """
    Execute a stored procedure that may return multiple result sets
    Returns two separate DataFrames (df1, df2)
    If second result set doesn't exist, returns empty DataFrame for df2
    """
    param_placeholders = ', '.join(['%s'] * len(params)) if params else ''
    query = f"CALL {procedure_name}({param_placeholders})"
    
    try:
        with connection.cursor() as cursor:
            cursor.execute(query, params or ())
            
            # Get first result set
            df1 = pd.DataFrame()
            df2 = pd.DataFrame()
            
            try:
                rows = cursor.fetchall()
                if rows:
                    columns = [col[0] for col in cursor.description]
                    df1 = pd.DataFrame(rows, columns=columns)
                    print(f"First result set shape: {df1.shape}")
                else:
                    print("First result set is empty")
            except Exception as e:
                print(f"Error fetching first result set: {e}")
                df1 = pd.DataFrame()
            
            # Try to get second result set
            try:
                if cursor.nextset():
                    rows = cursor.fetchall()
                    if rows:
                        columns = [col[0] for col in cursor.description]
                        df2 = pd.DataFrame(rows, columns=columns)
                        print(f"Second result set shape: {df2.shape}")
                    else:
                        print("Second result set is empty")
                else:
                    print("No second result set available")
            except Exception as e:
                # Handle case where no more result sets exist
                if "no result set" in str(e).lower() or "no data" in str(e).lower():
                    print("No second result set available")
                else:
                    print(f"Error fetching second result set: {e}")
                df2 = pd.DataFrame()
            
            return df1, df2
            
    except Exception as e:
        print(f"Error executing stored procedure {procedure_name}: {e}")
        # Return empty DataFrames on error
        return pd.DataFrame(), pd.DataFrame()
    
    
def create_pdf_Logo(df, output_path, title="Database Report", subtitle=None, 
               orientation='portrait', page_size='A4', include_chart=False, col_widths=None,logo_path=None):
    """
    Generate a PDF report from a pandas DataFrame
    
    Args:
        df: pandas DataFrame containing the data
        output_path: Path where PDF will be saved
        title: Report title
        subtitle: Report subtitle
        orientation: 'portrait' or 'landscape'
        page_size: 'A4' or 'letter'
        include_chart: Whether to include a simple chart
        col_widths: List of column widths for the table
    """
    # Set up PDF document properties
    if page_size == 'A4':
        page_size_tuple = A4
    else:  # letter
        page_size_tuple = letter
        
    if orientation == 'landscape':
        page_size_tuple = landscape(page_size_tuple)
    
    doc = SimpleDocTemplate(
        "media/ReportFiles/"+output_path, 
        pagesize=page_size_tuple,
        rightMargin=0.5 * inch,
        leftMargin=0.5 * inch,
        topMargin=0.5 * inch,
        bottomMargin=0.5 * inch
    )
    
    # Get styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(name='CenteredTitle', alignment=1, fontSize=16, spaceAfter=12)
    subtitle_style = ParagraphStyle(name='CenteredSubtitle', alignment=1, fontSize=12, spaceAfter=6)
    normal_style = styles['Normal']
    
    # Create document elements
    elements = []
    
    if logo_path:
        logo = Image(logo_path, width=1.5 * inch, height=0.75 * inch)  # Adjust width and height as needed
        logo.hAlign = 'CENTER'
        elements.append(logo)
        elements.append(Spacer(1, 0.25 * inch))

    # Add title and subtitle
    elements.append(Paragraph(title, title_style))
    if subtitle:
        elements.append(Paragraph(subtitle, subtitle_style))
    
    # Add timestamp
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    elements.append(Paragraph(f"Generated: {timestamp}", normal_style))
    elements.append(Spacer(1, 0.25 * inch))
    
    # Create chart if requested
    if include_chart and not df.empty and df.shape[0] > 0 and df.shape[1] > 1:
        try:
            numeric_cols = df.select_dtypes(include=['number']).columns
            if len(numeric_cols) > 0:
                plt.figure(figsize=(7, 4))
                
                if df.shape[0] <= 10:
                    df.plot(x=df.columns[0], y=numeric_cols[0], kind='bar')
                else:
                    df.plot(x=df.columns[0], y=numeric_cols[0], kind='line')
                    
                plt.title(f"Chart of {numeric_cols[0]}")
                plt.tight_layout()
                
                img_data = BytesIO()
                plt.savefig(img_data, format='png')
                img_data.seek(0)
                
                img = Image(img_data, width=5 * inch, height=2 * inch)
                elements.append(img)
                elements.append(Spacer(1, 0.2 * inch))
        except Exception as e:
            print(f"Error creating chart: {e}")
    
    # Add table with data
    if not df.empty:
        # Convert DataFrame to list of lists and wrap each cell in a Paragraph
        data = [[Paragraph(str(item), normal_style) for item in row] for row in [df.columns.tolist()] + df.values.tolist()]
        
        # Create table
        if col_widths is None:
            col_widths = [1.5 * inch] * len(df.columns)  # Default width if not provided
        
        # Calculate row heights based on wrapped text
        row_heights = []
        for row in data:
            max_height = 0
            for item in row:
                # Measure height of each paragraph after wrapping
                height = item.wrap(col_widths[row.index(item)], 0)[1]
                if height > max_height:
                    max_height = height
            row_heights.append(max_height + 10)  # Add padding to height

        # Create table with specified column widths and row heights
        table = Table(data, colWidths=col_widths, rowHeights=row_heights)
        
        # Add style
        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ])
        
        # Add alternating row colors
        for row in range(1, len(data)):
            if row % 2 == 0:
                style.add('BACKGROUND', (0, row), (-1, row), colors.lightgrey)
        
        table.setStyle(style)
        elements.append(table)
    else:
        elements.append(Paragraph("No data available", normal_style))
    
    # Build PDF
    doc.build(elements)
    print(f"PDF report saved to: {output_path}")
    return output_path

def create_pdf(df,dfHeaders, output_path, title="Database Report", subtitle=None,
               orientation='landscape', page_size='A4', include_chart=False):
    """
    Generate a PDF report from a pandas DataFrame

    Args:
        df: pandas DataFrame containing the data
        output_path: Path where PDF will be saved
        title: Report title
        subtitle: Report subtitle
        orientation: 'portrait' or 'landscape'
        page_size: 'A4' or 'letter'
        include_chart: Whether to include a simple chart
    """
    
    # Set up PDF document properties
    if page_size == 'A4':
        page_size_tuple = A4
    else:  # letter
        page_size_tuple = letter
        
    if orientation == 'landscape':
        page_size_tuple = landscape(page_size_tuple)

    doc = SimpleDocTemplate(
        "media/ReportFiles/" + output_path, 
        pagesize=page_size_tuple,
        rightMargin=0.5 * inch,
        leftMargin=0.5 * inch,
        topMargin=0.5 * inch,
        bottomMargin=0.5 * inch
    )

    # Get styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(name='CenteredTitle', alignment=1, fontSize=16, spaceAfter=12)
    subtitle_style = ParagraphStyle(name='CenteredSubtitle', alignment=1, fontSize=12, spaceAfter=6)
    normal_style = styles['Normal']

    # Create document elements
    elements = []

    logo_path = "Fincredit.png"
    if logo_path:
        logo = Image(logo_path, width=4.5 * inch, height=1.2 * inch)  # Adjust width and height as needed
        logo.hAlign = 'CENTER'
        elements.append(logo)
        elements.append(Spacer(1, 0.25 * inch))

    # Add title and subtitle
    elements.append(Paragraph(title, title_style))
    if subtitle:
        elements.append(Paragraph(subtitle, subtitle_style))

    # Add timestamp
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    elements.append(Paragraph(f"Generated: {timestamp}", normal_style))
    elements.append(Spacer(1, 0.25 * inch))

    # Handle dfHeaders - more flexible approach
    if not dfHeaders.empty:
        header_data = []
        
        # Since all items are in the same row, we need to transpose or iterate through columns
        print(f"dfHeaders shape: {dfHeaders.shape}")
        print(f"Number of rows: {len(dfHeaders)}")
        print(f"Number of columns: {len(dfHeaders.columns)}")
        
        # Check if we have a single row with multiple columns
        if len(dfHeaders) == 1:
            # Get the first (and only) row
            single_row = dfHeaders.iloc[0]
            print(f"Single row data: {single_row}")
            
            # Iterate through all columns in this row
            for col_name, value in single_row.items():
                if pd.notna(value) and str(value).strip():
                    header_data.append([Paragraph(f"<b>{col_name}:</b>", normal_style), Paragraph(str(value), normal_style)])
                    print(f"Added header: {col_name} = {value}")
        
        elif len(dfHeaders.columns) == 2 and len(dfHeaders) > 1:
            # If we have multiple rows with 2 columns (key-value format)
            key_col = dfHeaders.columns[0]
            value_col = dfHeaders.columns[1]
            
            for index, row in dfHeaders.iterrows():
                key = row.get(key_col, '')
                value = row.get(value_col, '')
                if pd.notna(key) and pd.notna(value) and str(key).strip() and str(value).strip():
                    header_data.append([Paragraph(f"<b>{key}:</b>", normal_style), Paragraph(str(value), normal_style)])
        
        else:
            # Fallback: display all data as is
            for index, row in dfHeaders.iterrows():
                for col_name, value in row.items():
                    if pd.notna(value) and str(value).strip():
                        header_data.append([Paragraph(f"<b>{col_name}:</b>", normal_style), Paragraph(str(value), normal_style)])
        
        if header_data:
            print(f"Created {len(header_data)} header items")
            header_table = Table(header_data, colWidths=[1.5 * inch, 5 * inch])
            
            header_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.whitesmoke),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ])
            header_table.setStyle(header_style)
            elements.append(header_table)
            elements.append(Spacer(1, 0.25 * inch))
        else:
            print("No valid header data found after processing dfHeaders")


    # Add table with data
    if not df.empty:
        # Add headers
        header_row = [Paragraph(str(col), styles['Heading4']) for col in df.columns if col != 'descr']
        
        # Processing rows
        data = []
        for index, row in df.iterrows():
            row_data = row.drop('descr', errors='ignore').tolist()  # Drop the 'desc' column
            
            # Check the desc column for actions
            if 'descr' in row and row['descr'] == 'PageBreak':
                if data:  # Only add table if there's data collected
                    # Create and style the table for the current data
                    table = Table(data)
                    apply_table_style(table)
                    elements.append(table)
                    elements.append(PageBreak())  # Add page break after the table
                # Start a new data collection for the next table
                data = [header_row] + [row_data]
            elif 'descr' in row and row['descr'] == 'Bold':
                # Add bold styled row data
                styled_row = [Paragraph( str(item), styles['Heading4']) for item in row_data]
                if not data:  # Initialize data with header if empty
                    data = [header_row]
                data.append(styled_row)
            else:
                # Regular row
                styled_row = [Paragraph(str(item), normal_style) for item in row_data]
                if not data:  # Initialize data with header if empty
                    data = [header_row]
                data.append(styled_row)

        # After processing all rows, check if there's any data to create a final table
        if data and len(data) > 1:  # At least one header row and one data row
            table = Table(data)
            apply_table_style(table)
            elements.append(table)
        else:
            elements.append(Paragraph("No data to display after processing.", normal_style))
    else:
        elements.append(Paragraph("No data available", normal_style))

    # Build PDF
    doc.build(elements)
    print(f"PDF report saved to: {output_path}")
    return output_path
def format_output(value):
    if isinstance(value, str):
        return f"String: {value}"
    elif isinstance(value, int):
        return f"Integer: {value:,}"  # Comma as thousands separator
    elif isinstance(value, float):
        return f"Float: {value:.2f}"  # 2 decimal places
    elif isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    else:
        return "Unsupported type"

def apply_table_style(table):
    """Apply styling to the table."""
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ])
    table.setStyle(style)


def parse_formatting_tags(text):
    """
    Parse formatting tags in text and return formatted result
    
    Supported tags:
    - [bold]text[/bold] - Makes text bold
    - [header]text[/header] - Makes text a header
    - [pagebreak] - Inserts a page break
    
    Args:
        text: Text with formatting tags
        
    Returns:
        List of elements with appropriate formatting
    """
    styles = getSampleStyleSheet()
    normal_style = styles['Normal']
    heading_style = styles['Heading2']
    
    # Create bold style
    bold_style = ParagraphStyle(
        'Bold', 
        parent=normal_style,
        fontName='Helvetica-Bold'
    )
    
    elements = []
    
    # Split by pagebreak tag
    parts = text.split('[pagebreak]')
    
    for i, part in enumerate(parts):
        # Process bold tags
        bold_parts = re.split(r'\[bold\](.*?)\[\/bold\]', part)
        
        for j, bold_part in enumerate(bold_parts):
            if j % 2 == 0:  # Regular text
                # Process header tags
                header_parts = re.split(r'\[header\](.*?)\[\/header\]', bold_part)
                
                for k, header_part in enumerate(header_parts):
                    if k % 2 == 0:  # Regular text
                        if header_part.strip():
                            elements.append(Paragraph(header_part, normal_style))
                    else:  # Header text
                        elements.append(Paragraph(header_part, heading_style))
            else:  # Bold text
                elements.append(Paragraph(bold_part, bold_style))
        
        # Add page break if not the last part
        if i < len(parts) - 1:
            elements.append(PageBreak())
    
    return elements

def create_formatted_pdf(formatted_text, output_path, title="Formatted Report", 
                        orientation='portrait', page_size='A4'):
    """
    Generate a PDF with formatted text
    
    Args:
        formatted_text: Text with formatting tags
        output_path: Path where PDF will be saved
        title: Report title
        orientation: 'portrait' or 'landscape'
        page_size: 'A4' or 'letter'
    """
    # Set up PDF document properties
    if page_size == 'A4':
        page_size_tuple = A4
    else:  # letter
        page_size_tuple = letter
        
    if orientation == 'landscape':
        page_size_tuple = landscape(page_size_tuple)
    
    doc = SimpleDocTemplate(
        output_path, 
        pagesize=page_size_tuple,
        rightMargin=0.5*inch,
        leftMargin=0.5*inch,
        topMargin=0.5*inch,
        bottomMargin=0.5*inch
    )
    
    # Get styles
    styles = getSampleStyleSheet()
    title_style = styles['Heading1']
    
    # Create document elements
    elements = []
    
    # Add title
    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 0.25*inch))
    
    # Add formatted content
    elements.extend(parse_formatting_tags(formatted_text))
    
    # Build PDF
    doc.build(elements)
    print(f"Formatted PDF report saved to: {output_path}")
    return output_path

@csrf_exempt
def RequestReportxlsx_view(request, *args, **kwargs):

    from openpyxl import Workbook
    from openpyxl.drawing.image import Image
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import Alignment, Font

    if Configs.is_ajax(request):
        try:
            hdnReportId = request.POST.get('hdnReportId', None)
            view_report = ReportFile.objects.filter(ReportId=hdnReportId).values().all()

            for setting in view_report.iterator():
                ReportName = setting['ReportName']         # Accessing using dictionary key
                ReportSDesc = setting['ReportDescription']  # Accessing using dictionary key
                ReportProc = setting['SPObjectName'] 
                

            form_data = {key: request.POST.get(key) for key in request.POST}
            search_json = json.dumps(form_data)
            params = construct_parameters(search_json)
            print(params)
            
            # MySQL connection details (Edit Accordingly)
            host = "localhost"  # Replace with your MySQL host
            port = 3306               # Replace with your MySQL port
            database = "fincredit"      # Replace with your database name
            username = "root"    # Replace with your username
            password = "Pass1"    # Replace with your password

            # MySQL connection details (Edit Accordingly)
            #host = "dev.pathwaystechnologies.com"  # Replace with your MySQL host
            #port = 7878               # Replace with your MySQL port
            #database = "apexbmblive"      # Replace with your database name
            #username = "apexcbs"    # Replace with your username apexcbs/v$Apex@2025
            #password = "v$Apex@2025"    # Replace with your password

            #Connect to database and execute query
            connection = connect_to_mysql(host, port, database, username, password)
            if connection:
                # Execute SQL query
                #df = execute_query(connection, example_query)
                
                # Alternative: Execute a stored procedure
                df,dfheaders = execute_stored_procedure(connection, ReportProc,params)
                #print(df)
                #print(dfheaders)
               # if 'descr' in df.columns:
               #     df = df.drop(columns=['descr'])
                    
                print('Executing Report')
                now = datetime.now()

                # Format the datetime string
                formatted_datetime = now.strftime("%Y%m%d_%H%M%S")
                

                # Generate EXCEL report
                if not df.empty:
                    # Create a new workbook
                    wb = Workbook()
                    ws = wb.active
                    
                    # Add a logo (replace 'logo.png' with your logo path)
                    logo = Image("Fincredit.png")  # Ensure the logo file exists
                    logo.height = 50  # Adjust height as needed
                    logo.width = 350  # Adjust width as needed
                    ws.add_image(logo, 'D1')  # Place logo in cell A1

                    # Add report title below the logo (row 2)
                    report_title = ReportName  # Replace with your title
                    ws['A5'] = report_title
                    ws['A5'].font = Font(size=16, bold=True)
                    ws.merge_cells('A5:H5')  # Merge cells for the title (adjust range as needed)
                    ws['A5'].alignment = Alignment(horizontal='center')
                   # for cell in ws[8]:
                   #     cell.font = Font(size=14, bold=True)
                    print(' headers')
                    Report_Filters = construct_header(search_json)

                    print('end headers')
                    #Report_Filters ="Report For Branch " + params[0] + " To Branch " + params[1] + " From " + params[2] + " To " + params[3]
                    #ws['A9'].font = Font(size=14, bold=True) 
                    #ws.merge_cells('A9:C9')
                    ws['B6'].alignment = Alignment(horizontal='center') 
                    ws['B6'] = Report_Filters
                    ws.merge_cells('A6:D6')
                    print(Report_Filters)
                    # Write DataFrame starting from row 10 (leaving space for logo and title)
                    ##for r in dataframe_to_rows(df, index=False, header=True):

                    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 7):
                        ws.append(row)
                        
                        descr_col_idx = None
                        if 'descr' in df.columns:
                            descr_col_idx = df.columns.get_loc('descr')
                        
                        current_row = ws.max_row
                        if descr_col_idx is not None and row[descr_col_idx] == 'Bold':
                            for cell in ws[current_row]:
                                cell.font = Font(bold=False)
                            
                        if descr_col_idx is not None:
                            # Since we excluded it from row_data, we need to adjust for 0-based vs 1-based indexing
                            excel_descr_col = descr_col_idx + 1  # Convert pandas 0-index to Excel 1-index
                            ws.column_dimensions[get_column_letter(excel_descr_col)].hidden = True
	
	                    # Format numeric cells (skip header row when r_idx=4)
                        if r_idx > 7:  # Data starts at row 5
                            for c_idx, value in enumerate(row, 1):
                                cell = ws.cell(row=r_idx, column=c_idx)
                                
                                # Check if numeric and format
                                if isinstance(value, (int, float)) and value is not None:
                                    cell.number_format = '#,##0'  # Base format
                                    if isinstance(value, float):
                                        cell.number_format = '#,##0.00' 




                    

                    for cell in ws[7]:
                        cell.font = Font(size=14, bold=True)
  

                        last_row = ws.max_row
                        for cell in ws[last_row]:
                            cell.font = Font(size=11, bold=False)

                    # Adjust column widths (optional)
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = (max_length) #* 1.2
                        ws.column_dimensions[column_letter].width = adjusted_width
                    
                    print("df here")    
                    # Save the file
                    output_path = os.path.join("media/ReportFiles", f"report_{formatted_datetime}.xlsx")  # Saves as 'media/report.xlsx'
                    wb.save(output_path)
                    print("done")
                print('Executing Report Completed')
                # Close connection
            connection.close()
            fileName = f"report_{formatted_datetime}.xlsx"

            response = {
                'details': 'Report Generated!',
                'file_url': request.build_absolute_uri(f'/media/ReportFiles/{fileName}')  # URL to download the file
            }
            
            return JsonResponse({"Success": response}, status=200)
        except  Exception as e:
            eString = str(e)
            print(eString)
            return JsonResponse({"Exception": eString}, status=404)
